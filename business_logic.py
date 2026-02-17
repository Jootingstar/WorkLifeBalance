import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

# 상수 정의
MAX_YEARS = 20
YEAR_COLUMNS = [f"{i+1}년" for i in range(MAX_YEARS)]
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def process_source_data(source_file, sheet_name):
    """소스 파일에서 데이터를 처리합니다."""
    df_source = pd.read_excel(source_file, sheet_name=sheet_name)
    
    # 컬럼명 설정
    if len(df_source.columns) >= 4:
        df_source.columns = ['name', 'start_date', 'end_date', 'days'] + list(df_source.columns[4:])
    
        # 병합된 셀 처리
        df_source['name'] = df_source['name'].ffill()
    
        # value 컬럼 제거 (처음 4개 컬럼만 사용)
        processed_data = df_source[['name', 'start_date', 'end_date', 'days']].copy()
    
        return df_source, processed_data
    else:
        return None, None


def load_teamSheet(target_path, sheet_name):
    """Team 시트를 읽어옵니다."""
    try:
        df_team = pd.read_excel(target_path, sheet_name=sheet_name)  # @paul.wbc
        
        if df_team.empty or len(df_team.columns) == 0:
            return pd.DataFrame(columns=['name', 'start_date', 'end_date', 'days']), False
        else:
            df_team.columns = ['name', 'start_date', 'end_date', 'days']
            return df_team, True
    except Exception as e:
        return pd.DataFrame(columns=['name', 'start_date', 'end_date', 'days']), False


def find_new_data(source_data, df_team):
    """중복을 제외한 새 데이터를 찾습니다.
    name + start_date + end_date + days가 모두 같은 빼면 중복으로 간주합니다."""
    if not df_team.empty:
        # name + start_date + end_date + days 기준으로 중복 체크
        merged = source_data.merge(
            df_team[['name', 'start_date', 'end_date', 'days']],
            on=['name', 'start_date', 'end_date', 'days'],
            how='left',
            indicator=True
        )        
        new_data = merged[merged['_merge'] == 'left_only'].drop('_merge', axis=1)
    else:
        new_data = source_data
    
    return new_data


def merge_and_sort_data(df_team, new_data):
    """기존 데이터와 새 데이터를 병합하고 정렬합니다.
    new_data에 없는 name + start_date와 중복되는 기존 데이터는 제거(영입으기) 합니다."""
    
    if not new_data.empty and not df_team.empty:
        # new_data의 (name, start_date) 조합 추출
        new_keys = new_data[['name', 'start_date']].copy()
        new_keys['_to_remove'] = True
        
        # 기존 데이터에서 중복되는 항목 제거
        df_team_with_flag = df_team.merge(new_keys, on=['name', 'start_date'], how='left')
        df_team_filtered = df_team_with_flag[df_team_with_flag['_to_remove'].isna()].drop('_to_remove', axis=1)
        
        # 중복 제거된 기존 데이터와 새 데이터 병합
        df_updated = pd.concat([df_team_filtered, new_data], ignore_index=True)
    elif not new_data.empty:
        df_updated = new_data
    else:
        df_updated = df_team
    
    df_updated = df_updated.sort_values(
        by=['name', 'start_date'],
        ascending=[True, False]  # name은 오름차순, start_date는 내림차순(최신이 위)
    )    
    return df_updated


def adjust_column_width(target_path, sheet_name):
    """지정된 시트의 셀 너비를 자동 조정합니다."""
    wb = openpyxl.load_workbook(target_path)
    
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"시트 '{sheet_name}'를 찾을 수 없습니다.")
    
    worksheet = wb[sheet_name]
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # 최소 너비 15, 최대 너비 50으로 제한하고 여백의 여유 추가
        adjusted_width = min(max(max_length + 2, 15), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(target_path)
    wb.close()


def save_to_excel(target_path, df_updated, sheet_name):
    """업데이트된 데이터를 Excel 파일에 저장합니다."""
    # 날짜를 date 형식으로 변환
    df_to_save = df_updated.copy()
    df_to_save['start_date'] = df_to_save['start_date'].dt.date
    df_to_save['end_date'] = df_to_save['end_date'].dt.date
    
    # Excel 파일에 쓰기 (기존 시트를 유지하면서 Team만 업데이트)
    with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 셀 너비 자동 조정
    adjust_column_width(target_path, sheet_name)


def create_yearly_summary(target_path):
    """2번째 시트부터 마지막 시트까지를 읽어와 '재현이네' 시트에 작성합니다."""
    wb = openpyxl.load_workbook(target_path)
    sheet_names = wb.sheetnames
    
    # 2번째 시트부터 마지막 시트까지 데이터 수집
    all_data = []
    sheet_name_mapping = {}  # 시트 순서 -> 시트명 매핑
    
    for sheet_idx, sheet_name in enumerate(sheet_names[1::], start=1):  # 첫 번째 시트 제외, 인덱스 추가
        sheet_name_mapping[sheet_idx] = sheet_name # 매핑 저장
        try:
            df = pd.read_excel(target_path, sheet_name=sheet_name)
            
            # 컬럼이 4개 이상이고 필요한 데이터가 있는 경우
            if len(df.columns) >= 4:
                df.columns = ['name', 'start_date', 'end_date', 'days'] + list(df.columns[4:])
                
                # 이름이 있는 행만 처리
                df = df[df['name'].notna()].copy()
                df = df[df['name'].astype(str).str.strip() != ''].copy()

                # 날짜에서 연도 추출
                df['start_date'] = pd.to_datetime(df['start_date'], errors='coerce')
                df['year'] = df['start_date'].dt.year
                
                # 필요한 컬럼만 선택 + 시트 순서 추가
                subset = df[['name', 'year', 'days']].copy()
                subset = subset[subset['year'].notna() & subset['days'].notna()]
                subset['sheet_order'] = sheet_idx  # 시트 순서 추가
                
                all_data.append(subset)
        except Exception as e:
            continue
    
    if not all_data:
        wb.close()
        return
    
    # 모든 데이터 합치기
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # name과 year별로 days 합계 계산
    summary = combined_df.groupby(['name', 'year', 'sheet_order'])['days'].sum().reset_index()
    
    # 각 name의 첫 번째 등장 시트 순서 구하기 (우선순위 결정용)
    name_priority = summary.groupby('name')['sheet_order'].min().reset_index()
    name_priority.columns = ['name', 'first_sheet']
    
    # 시트명 추가
    name_priority['sheet_name'] = name_priority['first_sheet'].map(sheet_name_mapping)
    
    # name과 year별로 days 합계 (시트 순서 제거)
    summary_agg = summary.groupby(['name', 'year'])['days'].sum().reset_index()
    
    # 피벗 테이블 생성: name을 행, year를 열로
    pivot_df = summary_agg.pivot(index='name', columns='year', values='days').fillna(0)
    
    # 연도를 "N년차" 형식으로 변환
    years = sorted(pivot_df.columns)
    year_mapping = {year: f"{i+1}년" for i, year in enumerate(years)}
    pivot_df.columns = [year_mapping[year] for year in years]
    pivot_df = pivot_df.reset_index()
    
    # 2여석치키 모든 열 추가 (없는 년차는 0으로)
    for year_col in YEAR_COLUMNS:
        if year_col not in pivot_df.columns:
            pivot_df[year_col] = 0
    
    # name별 첫 등장 시트 순서와 시트명 병합
    pivot_df = pivot_df.merge(name_priority[['name', 'first_sheet', 'sheet_name']], on='name', how='left')
    
    # 시트 순서 인지, 그 다음 name으로 정렬
    pivot_df = pivot_df.sort_values(['first_sheet', 'name'])
    pivot_df = pivot_df.drop('first_sheet', axis=1)  # 정렬 후 제거
    
    # sheet_name, name, 그리고 1년차부터 2여석치까지 순서대로 컬럼 정렬
    pivot_df = pivot_df[['sheet_name', 'name'] + YEAR_COLUMNS]
    
    # '재현이네' 시트에 쓰기
    if '재현이네' in wb.sheetnames:
        del wb['재현이네']
    
    ws = wb.create_sheet('재현이네', 0)  # 첫 번째 위치에 생성
    
    # 1행: A~여치 병합하고 "재현이네" 제목, I열부터 1년차~20년차 헤더
    ws.merge_cells('A1:H1')
    ws['A1'] = '재현이네'
    ws['A1'].alignment = CENTER_ALIGN
    for col_idx, year_label in enumerate(YEAR_COLUMNS, start=9):
        cell = ws.cell(row=1, column=col_idx, value=year_label)
        cell.alignment = CENTER_ALIGN
    
    # 2행: A열=team, B열=name, I열부터 각 년차별 숫자
    ws['A2'] = 'team'
    ws['B2'] = 'name'
    # 1년차: 11개, 2년차부터: 15개 시작하여 2년마다 1개씩 증가
    for year_num in range(1, MAX_YEARS + 1):
        value = 11 if year_num == 1 else 15 + (year_num - 2) // 2
        cell = ws.cell(row=2, column=8 + year_num, value=f"{value}개")
        cell.alignment = CENTER_ALIGN
    

    # 데이터 작성 - 3행부터 시작
    for row_idx, row in enumerate(pivot_df.itertuples(index=False), start=3):
        ws.cell(row=row_idx, column=1, value=row[0])  # A열: 시트명
        ws.cell(row=row_idx, column=2, value=row[1])  # B열: name
        # I열부터 20년차까지 데이터
        for col_idx, value in enumerate(row[2:2 + MAX_YEARS], start=9):
            if value > 0:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CENTER_ALIGN
    
    wb.save(target_path)
    wb.close()